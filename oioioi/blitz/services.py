import posixpath
import re
import zipfile
from dataclasses import dataclass
from xml.etree import ElementTree

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils.translation import gettext_lazy as _
from unidecode import unidecode

from oioioi.base.utils import uploaded_file_name
from oioioi.blitz.models import BlitzContestConfig
from oioioi.contests.models import (
    Contest,
    ContestAttachment,
    ContestLink,
    ContestPermission,
    FilesMessage,
    LimitsVisibilityConfig,
    ProblemInstance,
    ProblemStatementConfig,
    RankingVisibilityConfig,
    RegistrationAvailabilityConfig,
    Round,
    SubmissionMessage,
    SubmitMessage,
    SubmissionsMessage,
)
from oioioi.participants.models import Participant, TermsAcceptedPhrase
from oioioi.problems.utils import copy_problem_instance
from oioioi.programs.models import ProgramsConfig


MATCH_CODE_SLUG_RE = re.compile(r"[^a-z0-9_-]+")
REQUIRED_COLUMNS = ("match_code", "player1_username", "player2_username")
CONTEST_ID_MAX_LENGTH = Contest._meta.get_field("id").max_length
CONTEST_NAME_MAX_LENGTH = Contest._meta.get_field("name").max_length
XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XLSX_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass(frozen=True)
class MatchDefinition:
    row_number: int
    match_code: str
    normalized_match_code: str
    contest_id: str
    contest_name: str
    player1: User
    player2: User


def normalize_match_code(value):
    normalized = unidecode(str(value or "")).lower().strip()
    normalized = MATCH_CODE_SLUG_RE.sub("-", normalized).strip("-_")
    return normalized


def _clean_cell_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _copy_contest_one_to_one(model, source_contest, target_contest):
    try:
        source_object = model.objects.get(contest=source_contest)
    except model.DoesNotExist:
        return

    target_object, _created = model.objects.get_or_create(contest=target_contest)
    update_fields = []
    for field in source_object._meta.concrete_fields:
        if field.primary_key or field.name == "contest":
            continue
        setattr(target_object, field.name, getattr(source_object, field.name))
        update_fields.append(field.name)
    if update_fields:
        target_object.save(update_fields=update_fields)


def _worksheet_column_index(cell_reference):
    letters = "".join(character for character in cell_reference if character.isalpha())
    index = 0
    for letter in letters.upper():
        index = index * 26 + ord(letter) - 64
    return index


def _extract_inline_text(element):
    return "".join(text or "" for text in element.itertext())


def _read_xlsx_rows_with_stdlib(workbook_file):
    with uploaded_file_name(workbook_file) as filename:
        with zipfile.ZipFile(filename) as archive:
            workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            first_sheet = workbook_root.find(f"{{{XLSX_MAIN_NS}}}sheets/{{{XLSX_MAIN_NS}}}sheet")
            if first_sheet is None:
                return []

            relationship_id = first_sheet.attrib.get(f"{{{XLSX_REL_NS}}}id")
            if not relationship_id:
                raise ValidationError(_("The workbook could not be read."))

            relationships_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            worksheet_target = None
            for relationship in relationships_root.findall(f"{{{XLSX_PACKAGE_REL_NS}}}Relationship"):
                if relationship.attrib.get("Id") == relationship_id:
                    worksheet_target = relationship.attrib.get("Target")
                    break
            if worksheet_target is None:
                raise ValidationError(_("The workbook could not be read."))

            worksheet_path = posixpath.normpath(posixpath.join("xl", worksheet_target))
            shared_strings = []
            try:
                shared_strings_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            except KeyError:
                pass
            else:
                shared_strings = [
                    "".join(text_node.text or "" for text_node in item.findall(f".//{{{XLSX_MAIN_NS}}}t"))
                    for item in shared_strings_root.findall(f"{{{XLSX_MAIN_NS}}}si")
                ]

            worksheet_root = ElementTree.fromstring(archive.read(worksheet_path))
            rows = []
            for row in worksheet_root.findall(f".//{{{XLSX_MAIN_NS}}}sheetData/{{{XLSX_MAIN_NS}}}row"):
                values = []
                current_column = 1
                for cell in row.findall(f"{{{XLSX_MAIN_NS}}}c"):
                    reference = cell.attrib.get("r", "")
                    column_index = _worksheet_column_index(reference) if reference else current_column
                    while current_column < column_index:
                        values.append("")
                        current_column += 1

                    cell_type = cell.attrib.get("t")
                    if cell_type == "inlineStr":
                        value = _extract_inline_text(cell)
                    elif cell_type == "s":
                        shared_string_index = cell.findtext(f"{{{XLSX_MAIN_NS}}}v")
                        if shared_string_index in (None, ""):
                            value = ""
                        else:
                            value = shared_strings[int(shared_string_index)]
                    else:
                        value = cell.findtext(f"{{{XLSX_MAIN_NS}}}v", default="")

                    values.append(value)
                    current_column += 1
                rows.append(tuple(values))
    return rows


def _read_xlsx_rows_with_openpyxl(workbook_file):
    from openpyxl import load_workbook

    with uploaded_file_name(workbook_file) as filename:
        workbook = load_workbook(filename=filename, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            return [tuple(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()


def _read_xlsx_rows(workbook_file):
    try:
        return _read_xlsx_rows_with_openpyxl(workbook_file)
    except ImportError:
        return _read_xlsx_rows_with_stdlib(workbook_file)


def _parse_match_definitions(source_contest, level_label, contest_id_prefix, workbook_file):
    errors = []
    raw_rows = []

    try:
        rows = _read_xlsx_rows(workbook_file)
        header = rows[0] if rows else None
        if header is None:
            raise ValidationError(_("The workbook is empty."))

        normalized_header = tuple(_clean_cell_value(cell).lower() for cell in header[: len(REQUIRED_COLUMNS)])
        if normalized_header != REQUIRED_COLUMNS:
            raise ValidationError(
                _("Missing header or invalid columns: %(header)s. Expected: %(expected)s.")
                % {
                    "header": ", ".join(_clean_cell_value(cell) for cell in header[: len(REQUIRED_COLUMNS)]),
                    "expected": ", ".join(REQUIRED_COLUMNS),
                }
            )

        for row_number, row in enumerate(rows[1:], start=2):
            values = [_clean_cell_value(cell) for cell in row[: len(REQUIRED_COLUMNS)]]
            while len(values) < len(REQUIRED_COLUMNS):
                values.append("")
            if not any(values):
                continue
            raw_rows.append((row_number, *values))
    except ValidationError:
        raise
    except Exception as e:
        raise ValidationError(_("The workbook could not be read.")) from e

    if not raw_rows:
        raise ValidationError(_("The workbook does not contain any matches."))

    usernames = set()
    normalized_codes = {}
    contest_ids = {}
    seen_players = {}

    for row_number, match_code, player1_username, player2_username in raw_rows:
        if not match_code:
            errors.append(_("Row %(row_number)d: match_code is required.") % {"row_number": row_number})
        if not player1_username:
            errors.append(_("Row %(row_number)d: player1_username is required.") % {"row_number": row_number})
        if not player2_username:
            errors.append(_("Row %(row_number)d: player2_username is required.") % {"row_number": row_number})
        if player1_username and player2_username and player1_username == player2_username:
            errors.append(_("Row %(row_number)d: both players must be different.") % {"row_number": row_number})

        for username in (player1_username, player2_username):
            if not username:
                continue
            usernames.add(username)
            previous_row = seen_players.setdefault(username, row_number)
            if previous_row != row_number:
                errors.append(
                    _("Rows %(first_row)d and %(second_row)d: %(username)s is scheduled in more than one match.")
                    % {
                        "first_row": previous_row,
                        "second_row": row_number,
                        "username": username,
                    }
                )

        normalized_match_code = normalize_match_code(match_code)
        if match_code and not normalized_match_code:
            errors.append(_("Row %(row_number)d: %(match_code)s cannot be normalized to a valid match code.") % {"row_number": row_number, "match_code": match_code})
            continue

        if normalized_match_code:
            previous_row = normalized_codes.setdefault(normalized_match_code, row_number)
            if previous_row != row_number:
                errors.append(
                    _("Rows %(first_row)d and %(second_row)d: match codes collide after normalization (%(match_code)s).")
                    % {
                        "first_row": previous_row,
                        "second_row": row_number,
                        "match_code": normalized_match_code,
                    }
                )

            contest_id = f"{contest_id_prefix}-{normalized_match_code}"
            if len(contest_id) > CONTEST_ID_MAX_LENGTH:
                errors.append(
                    _("Row %(row_number)d: generated contest ID %(contest_id)s is longer than %(max_length)d characters.")
                    % {
                        "row_number": row_number,
                        "contest_id": contest_id,
                        "max_length": CONTEST_ID_MAX_LENGTH,
                    }
                )
            previous_row = contest_ids.setdefault(contest_id, row_number)
            if previous_row != row_number:
                errors.append(
                    _("Rows %(first_row)d and %(second_row)d: generated contest ID %(contest_id)s is duplicated.")
                    % {
                        "first_row": previous_row,
                        "second_row": row_number,
                        "contest_id": contest_id,
                    }
                )

    existing_users = User.objects.in_bulk(usernames, field_name="username") if usernames else {}
    for username in sorted(usernames - set(existing_users)):
        errors.append(_("Row %(row_number)d: user %(username)s does not exist.") % {"row_number": seen_players[username], "username": username})

    for contest_id in sorted(Contest.objects.filter(id__in=contest_ids).values_list("id", flat=True)):
        errors.append(_("Generated contest ID %(contest_id)s already exists.") % {"contest_id": contest_id})

    if errors:
        raise ValidationError(errors)

    definitions = []
    for row_number, match_code, player1_username, player2_username in raw_rows:
        normalized_match_code = normalize_match_code(match_code)
        contest_id = f"{contest_id_prefix}-{normalized_match_code}"
        contest_name = _("%(source)s - %(level)s - %(match_code)s: %(player1)s vs %(player2)s") % {
            "source": source_contest.name,
            "level": level_label,
            "match_code": match_code,
            "player1": player1_username,
            "player2": player2_username,
        }
        if len(contest_name) > CONTEST_NAME_MAX_LENGTH:
            raise ValidationError(
                _("Row %(row_number)d: generated contest name is longer than %(max_length)d characters.")
                % {
                    "row_number": row_number,
                    "max_length": CONTEST_NAME_MAX_LENGTH,
                }
            )
        definitions.append(
            MatchDefinition(
                row_number=row_number,
                match_code=match_code,
                normalized_match_code=normalized_match_code,
                contest_id=contest_id,
                contest_name=contest_name,
                player1=existing_users[player1_username],
                player2=existing_users[player2_username],
            )
        )

    return definitions


def _clone_rounds(source_contest, target_contest):
    round_map = {}
    for source_round in Round.objects.filter(contest=source_contest).order_by("start_date", "id"):
        target_round = Round.objects.create(
            contest=target_contest,
            name=source_round.name,
            start_date=source_round.start_date,
            end_date=source_round.end_date,
            results_date=source_round.results_date,
            public_results_date=source_round.public_results_date,
            is_trial=source_round.is_trial,
        )
        round_map[source_round.id] = target_round
    return round_map


def _clone_problem_instances(source_contest, target_contest, round_map):
    queryset = ProblemInstance.objects.filter(contest=source_contest).select_related("problem", "round").order_by("round__start_date", "order", "id")
    for source_problem_instance in queryset:
        source_round_id = source_problem_instance.round_id
        source_short_name = source_problem_instance.short_name
        source_submissions_limit = source_problem_instance.submissions_limit
        source_order = source_problem_instance.order
        source_needs_rejudge = source_problem_instance.needs_rejudge

        target_problem_instance = copy_problem_instance(source_problem_instance, target_contest)
        target_problem_instance.round = round_map.get(source_round_id)
        target_problem_instance.short_name = source_short_name
        target_problem_instance.submissions_limit = source_submissions_limit
        target_problem_instance.order = source_order
        target_problem_instance.needs_rejudge = source_needs_rejudge
        target_problem_instance.save()


def _clone_attachments(source_contest, target_contest, round_map):
    for source_attachment in ContestAttachment.objects.filter(contest=source_contest).select_related("round").order_by("id"):
        source_attachment.content.open("rb")
        try:
            attachment_content = ContentFile(source_attachment.content.read(), name=source_attachment.download_name)
        finally:
            source_attachment.content.close()

        target_attachment = ContestAttachment(
            contest=target_contest,
            description=source_attachment.description,
            round=round_map.get(source_attachment.round_id),
            pub_date=source_attachment.pub_date,
        )
        target_attachment.content.save(source_attachment.download_name, attachment_content, save=False)
        target_attachment.save()


def _clone_links(source_contest, target_contest):
    for source_link in ContestLink.objects.filter(contest=source_contest).order_by("id"):
        ContestLink.objects.create(
            contest=target_contest,
            description=source_link.description,
            url=source_link.url,
            order=source_link.order,
        )


def _clone_permissions(source_contest, target_contest):
    permissions = [
        ContestPermission(user=permission.user, contest=target_contest, permission=permission.permission)
        for permission in ContestPermission.objects.filter(contest=source_contest).select_related("user")
    ]
    if permissions:
        ContestPermission.objects.bulk_create(permissions)


def _clone_contest_configs(source_contest, target_contest):
    for model in (
        BlitzContestConfig,
        ProgramsConfig,
        TermsAcceptedPhrase,
        ProblemStatementConfig,
        RankingVisibilityConfig,
        LimitsVisibilityConfig,
        RegistrationAvailabilityConfig,
        FilesMessage,
        SubmissionsMessage,
        SubmitMessage,
        SubmissionMessage,
    ):
        _copy_contest_one_to_one(model, source_contest, target_contest)


def _assign_match_participants(contest, match_definition):
    Participant.objects.bulk_create(
        [
            Participant(contest=contest, user=match_definition.player1, status="ACTIVE"),
            Participant(contest=contest, user=match_definition.player2, status="ACTIVE"),
        ]
    )


def _clone_contest_for_match(source_contest, match_definition):
    target_contest = Contest.objects.create(
        id=match_definition.contest_id,
        name=match_definition.contest_name,
        controller_name=source_contest.controller_name,
        default_submissions_limit=source_contest.default_submissions_limit,
        contact_email=source_contest.contact_email,
        judging_priority=source_contest.judging_priority,
        judging_weight=source_contest.judging_weight,
        enable_editor=source_contest.enable_editor,
        show_contest_rules=source_contest.show_contest_rules,
        school_year=source_contest.school_year,
        is_archived=False,
    )

    round_map = _clone_rounds(source_contest, target_contest)
    _clone_problem_instances(source_contest, target_contest, round_map)
    _clone_contest_configs(source_contest, target_contest)
    _clone_links(source_contest, target_contest)
    _clone_attachments(source_contest, target_contest, round_map)
    _clone_permissions(source_contest, target_contest)
    _assign_match_participants(target_contest, match_definition)

    target_contest.controller.adjust_contest()
    return target_contest


def generate_match_contests(source_contest, level_label, contest_id_prefix, workbook_file):
    match_definitions = _parse_match_definitions(
        source_contest=source_contest,
        level_label=level_label,
        contest_id_prefix=contest_id_prefix,
        workbook_file=workbook_file,
    )

    with transaction.atomic():
        created_contests = [_clone_contest_for_match(source_contest, match_definition) for match_definition in match_definitions]

    return created_contests
